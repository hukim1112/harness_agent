"""
===============================================================================
[AAWS Tools] Analyst — 데이터 분석·시각화·보고서 전용 도구 6종
===============================================================================
data_profiler   : 자동 통계 프로파일링 (pandas 기반)
data_query      : pandas 코드 기반 분석 쿼리 실행
chart_generator : matplotlib/plotly 차트 생성
file_converter  : 데이터 포맷 변환 (CSV/JSON/Excel/Parquet)
excel_writer    : 수식/서식 기반 전문 Excel 보고서 생성
html_report     : 인터랙티브 HTML 대시보드 생성

📌 상세 노하우 참조: skills/analyst/ 디렉토리
   - xlsx_guide.md, chart_patterns.md, design_tokens.md, data_analysis.md
===============================================================================
"""

import os
import json
import time
import traceback
from pathlib import Path
from typing import Optional
from langchain_core.tools import tool
from pydantic import BaseModel, Field


# =============================================================================
# 유틸: 데이터 로딩 (확장자 자동 감지)
# =============================================================================

def _load_dataframe(file_path: str):
    """확장자 기반 자동 감지 데이터 로딩. pandas DataFrame 반환."""
    import pandas as pd

    abs_path = os.path.abspath(os.path.expanduser(file_path))
    if not os.path.exists(abs_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {abs_path}")

    ext = Path(abs_path).suffix.lower()

    if ext in ('.csv', '.tsv'):
        sep = '\t' if ext == '.tsv' else ','
        return pd.read_csv(abs_path, sep=sep)
    elif ext == '.json':
        return _load_json_smart(abs_path)
    elif ext == '.jsonl':
        return pd.read_json(abs_path, lines=True)
    elif ext in ('.xlsx', '.xls'):
        return pd.read_excel(abs_path)
    elif ext == '.parquet':
        return pd.read_parquet(abs_path)
    else:
        raise ValueError(f"지원하지 않는 파일 형식: {ext} (지원: csv, tsv, json, jsonl, xlsx, xls, parquet)")


def _load_json_smart(file_path: str):
    """JSON 구조 자동 감지: 배열/단일객체/중첩 리스트 필드"""
    import pandas as pd

    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if isinstance(data, list):
        return pd.DataFrame(data)
    elif isinstance(data, dict):
        # 최상위 키 중 리스트(of dicts) 값을 찾아 사용
        for key, val in data.items():
            if isinstance(val, list) and len(val) > 0 and isinstance(val[0], dict):
                return pd.DataFrame(val)
        # 단일 객체 → 1행 DataFrame
        return pd.DataFrame([data])
    else:
        raise ValueError("JSON 구조를 DataFrame으로 변환할 수 없습니다.")


def _truncate(text: str, max_len: int = 6000) -> str:
    """출력 길이 제한"""
    if len(text) > max_len:
        return text[:max_len] + f"\n... [TRUNCATED {len(text)-max_len} chars]"
    return text


# =============================================================================
# Tool 1: data_profiler — 자동 통계 프로파일링
# =============================================================================

class DataProfilerInput(BaseModel):
    file_path: str = Field(description="분석할 데이터 파일 경로 (JSON, CSV, Excel, Parquet 등)")
    sample_rows: int = Field(default=10, description="미리보기로 표시할 행 수. 기본값 10.")

@tool(args_schema=DataProfilerInput)
def data_profiler(file_path: str, sample_rows: int = 10) -> str:
    """데이터 파일을 읽어 자동으로 종합 통계 프로파일을 생성합니다.

    스키마(컬럼명·타입), 행 수, 기술 통계, 결측값 비율, 범주형 최빈값, 샘플 데이터를 한 번에 반환합니다.
    지원 형식: JSON, CSV, TSV, Excel (.xlsx/.xls), Parquet, JSONL

    ⚠️ 상세 분석 패턴이 필요하면: file_read('skills/analyst/data_analysis.md')

    Args:
        file_path: 분석할 데이터 파일 경로.
        sample_rows: 미리보기로 표시할 행 수. 기본값 10.

    Returns:
        마크다운 형식의 종합 통계 프로파일 문자열.
    """
    try:
        import pandas as pd
        df = _load_dataframe(file_path)

        sections = []
        abs_path = os.path.abspath(os.path.expanduser(file_path))

        # 1. 기본 정보
        sections.append(
            f"## 📋 기본 정보\n"
            f"- **파일**: `{abs_path}`\n"
            f"- **행 수**: {len(df):,}\n"
            f"- **열 수**: {len(df.columns)}\n"
            f"- **메모리**: {df.memory_usage(deep=True).sum() / 1024:.1f} KB"
        )

        # 2. 스키마 (unhashable 객체 안전 처리)
        schema_rows = []
        for col in df.columns:
            dtype = str(df[col].dtype)
            non_null = df[col].notna().sum()
            null_pct = df[col].isna().mean() * 100
            try:
                unique = df[col].nunique()
            except TypeError:
                # dict / list 등 unhashable 컬럼 대응
                unique = df[col].astype(str).nunique()
            schema_rows.append(f"| {col} | {dtype} | {non_null:,} | {null_pct:.1f}% | {unique:,} |")

        sections.append(
            "## 📊 스키마\n"
            "| 컬럼 | 타입 | 유효값 | 결측률 | 고유값 |\n"
            "|-------|------|--------|--------|--------|\n"
            + "\n".join(schema_rows)
        )

        # 3. 수치형 통계
        numeric_df = df.select_dtypes(include='number')
        if not numeric_df.empty:
            desc = numeric_df.describe().round(2)
            sections.append(f"## 📈 수치형 통계\n```\n{desc.to_string()}\n```")

        # 4. 범주형 최빈값 (unhashable 객체 안전 처리)
        cat_df = df.select_dtypes(include=['object', 'category', 'bool'])
        if not cat_df.empty:
            cat_rows = []
            for col in cat_df.columns:
                try:
                    top = df[col].value_counts().head(3)
                except TypeError:
                    top = df[col].astype(str).value_counts().head(3)
                top_str = ", ".join(f"{v}({c})" for v, c in top.items())
                cat_rows.append(f"| {col} | {top_str} |")
            sections.append(
                "## 🏷️ 범주형 최빈값 (상위 3)\n"
                "| 컬럼 | 상위 값(건수) |\n|-------|---------------|\n"
                + "\n".join(cat_rows)
            )

        # 5. 샘플 데이터
        sample = df.head(sample_rows)
        sections.append(f"## 🔍 샘플 데이터 (상위 {sample_rows}행)\n```\n{sample.to_string(max_colwidth=40)}\n```")

        return _truncate("\n\n".join(sections))

    except Exception as e:
        return f"DataProfiler 오류: {type(e).__name__}: {str(e)}\n{traceback.format_exc()[-400:]}"


# =============================================================================
# Tool 2: data_query — pandas 기반 분석 쿼리 실행
# =============================================================================

class DataQueryInput(BaseModel):
    file_path: str = Field(description="분석할 데이터 파일 경로")
    query_code: str = Field(description="실행할 pandas 코드. 변수 'df'로 데이터프레임에 접근합니다. 마지막 줄의 결과가 반환됩니다.")
    output_format: str = Field(default="table", description="결과 포맷: 'table' (마크다운 표), 'json' (JSON), 'stats' (기술 통계)")

@tool(args_schema=DataQueryInput)
def data_query(file_path: str, query_code: str, output_format: str = "table") -> str:
    """데이터 파일을 로드한 뒤 pandas 코드를 실행하여 분석 결과를 반환합니다.

    변수 'df'로 DataFrame에 접근하고, 'result' 변수에 결과를 저장하세요.
    result를 명시하지 않으면 마지막 표현식의 결과를 반환합니다.

    예시 query_code:
      "result = df.groupby('connection_type')['price'].agg(['mean', 'min', 'max', 'count'])"
      "result = df[df['price'] > 100000].sort_values('price', ascending=False)"

    ⚠️ 분석 패턴 참조: file_read('skills/analyst/data_analysis.md')

    Args:
        file_path: 데이터 파일 경로.
        query_code: 실행할 pandas 코드 문자열.
        output_format: 'table', 'json', 'stats' 중 하나.

    Returns:
        분석 결과 문자열 (지정 포맷).
    """
    try:
        import pandas as pd
        import numpy as np
        df = _load_dataframe(file_path)
    except Exception as e:
        return f"DataQuery 로드 오류: {str(e)}"

    # 안전한 실행 네임스페이스
    exec_globals = {
        '__builtins__': {
            'len': len, 'range': range, 'enumerate': enumerate, 'zip': zip,
            'sorted': sorted, 'reversed': reversed, 'list': list, 'dict': dict,
            'set': set, 'tuple': tuple, 'str': str, 'int': int, 'float': float,
            'bool': bool, 'min': min, 'max': max, 'sum': sum, 'abs': abs,
            'round': round, 'isinstance': isinstance, 'type': type, 'print': print,
            'True': True, 'False': False, 'None': None,
        },
        'pd': pd,
        'np': np,
        'df': df,
    }
    exec_locals = {}

    try:
        exec(query_code, exec_globals, exec_locals)
        result = exec_locals.get('result', None)

        # result가 없으면 마지막 줄 eval 시도
        if result is None:
            last_line = query_code.strip().split('\n')[-1].strip()
            if not last_line.startswith(('import ', 'from ', '#')):
                try:
                    result = eval(last_line, exec_globals, exec_locals)
                except Exception:
                    pass

        if result is None:
            return "쿼리 실행 완료. 'result' 변수에 결과를 할당하거나 마지막 줄에 표현식을 작성하세요."

        # 결과 포맷팅
        if isinstance(result, pd.DataFrame):
            if output_format == "json":
                return _truncate(result.to_json(orient='records', force_ascii=False, indent=2))
            elif output_format == "stats":
                return _truncate(f"```\n{result.describe().to_string()}\n```")
            else:
                return _truncate(f"({len(result)}행 × {len(result.columns)}열)\n```\n{result.to_string(max_colwidth=40)}\n```")
        elif isinstance(result, pd.Series):
            return _truncate(f"```\n{result.to_string()}\n```")
        else:
            return _truncate(str(result))

    except Exception as e:
        return f"DataQuery 실행 오류: {type(e).__name__}: {str(e)}\n{traceback.format_exc()[-500:]}"


# =============================================================================
# Tool 3: chart_generator — matplotlib/plotly 차트 생성
# =============================================================================

class ChartGeneratorInput(BaseModel):
    file_path: str = Field(description="데이터 파일 경로")
    chart_code: str = Field(description="차트 생성 코드. 'df'로 데이터 접근, 'fig'에 matplotlib Figure를 저장하거나 plotly Figure를 저장합니다.")
    output_path: str = Field(description="차트 저장 경로 (.png → matplotlib 이미지, .html → plotly 인터랙티브)")

@tool(args_schema=ChartGeneratorInput)
def chart_generator(file_path: str, chart_code: str, output_path: str) -> str:
    """데이터 파일에서 전문적인 차트를 생성하여 이미지/HTML로 저장합니다.

    matplotlib (PNG 이미지) 또는 plotly (HTML 인터랙티브) 차트를 생성합니다.
    코드에서 'df'로 DataFrame에 접근하고, 'fig'에 Figure를 저장하세요.

    matplotlib 예시:
      fig, ax = plt.subplots(figsize=(12, 6))
      df.groupby('category')['value'].mean().plot(kind='barh', ax=ax)
      ax.set_title('카테고리별 평균값')

    plotly 예시:
      import plotly.express as px
      fig = px.bar(df, x='category', y='value', title='비교')

    ⚠️ 차트 패턴 참조: file_read('skills/analyst/chart_patterns.md')

    Args:
        file_path: 데이터 파일 경로.
        chart_code: matplotlib/plotly 차트 생성 코드.
        output_path: 저장 경로 (.png 또는 .html).

    Returns:
        생성된 차트 파일 경로와 메타 정보.
    """
    try:
        import pandas as pd
        import numpy as np
        df = _load_dataframe(file_path)
    except Exception as e:
        return f"ChartGenerator 로드 오류: {str(e)}"

    abs_output = os.path.abspath(os.path.expanduser(output_path))
    os.makedirs(os.path.dirname(abs_output), exist_ok=True)
    ext = Path(abs_output).suffix.lower()

    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt

        # 한국어 폰트 설정 (크로스 플랫폼)
        import platform
        if platform.system() == 'Windows':
            plt.rcParams['font.family'] = 'Malgun Gothic'
        else:
            # Linux(WSL) — NanumGothic 우선, 없으면 DejaVu Sans
            from matplotlib.font_manager import fontManager
            available = {f.name for f in fontManager.ttflist}
            if 'NanumGothic' in available:
                plt.rcParams['font.family'] = 'NanumGothic'
            elif 'NanumBarunGothic' in available:
                plt.rcParams['font.family'] = 'NanumBarunGothic'
            else:
                plt.rcParams['font.family'] = 'DejaVu Sans'
        plt.rcParams['axes.unicode_minus'] = False
        plt.rcParams['figure.dpi'] = 150

        exec_globals = {
            '__builtins__': {
                'len': len, 'range': range, 'enumerate': enumerate, 'zip': zip,
                'sorted': sorted, 'list': list, 'dict': dict, 'str': str,
                'int': int, 'float': float, 'min': min, 'max': max,
                'sum': sum, 'abs': abs, 'round': round, 'print': print,
                'True': True, 'False': False, 'None': None,
            },
            'pd': pd,
            'np': np,
            'df': df,
            'plt': plt,
            'matplotlib': matplotlib,
        }

        # plotly 사용 가능 시 추가
        try:
            import plotly.express as px
            import plotly.graph_objects as go
            exec_globals['px'] = px
            exec_globals['go'] = go
        except ImportError:
            pass

        # seaborn 사용 가능 시 추가
        try:
            import seaborn as sns
            exec_globals['sns'] = sns
        except ImportError:
            pass

        exec_locals = {}
        exec(chart_code, exec_globals, exec_locals)

        fig = exec_locals.get('fig', None)

        if fig is None:
            # plt.gcf() 시도
            fig = plt.gcf()
            if not fig.axes:
                return "ChartGenerator 오류: 'fig' 변수에 Figure를 저장하거나, plt 함수로 차트를 생성하세요."

        # 저장
        if ext == '.html':
            # plotly Figure인 경우
            if hasattr(fig, 'write_html'):
                fig.write_html(abs_output, include_plotlyjs='cdn')
            else:
                return "ChartGenerator 오류: .html 출력에는 plotly Figure가 필요합니다."
        else:
            # matplotlib Figure
            if hasattr(fig, 'savefig'):
                fig.savefig(abs_output, dpi=150, bbox_inches='tight',
                            facecolor='white', edgecolor='none')
            else:
                plt.savefig(abs_output, dpi=150, bbox_inches='tight',
                            facecolor='white', edgecolor='none')
        plt.close('all')

        file_size = os.path.getsize(abs_output)
        return (
            f"✅ 차트 생성 완료\n"
            f"- **파일**: `{abs_output}`\n"
            f"- **크기**: {file_size/1024:.1f} KB\n"
            f"- **형식**: {ext.upper()}\n"
            f"- **데이터**: {len(df)}행 from `{file_path}`"
        )

    except Exception as e:
        plt.close('all')
        return f"ChartGenerator 실행 오류: {type(e).__name__}: {str(e)}\n{traceback.format_exc()[-500:]}"


# =============================================================================
# Tool 4: file_converter — 데이터 포맷 변환
# =============================================================================

class FileConverterInput(BaseModel):
    input_path: str = Field(description="입력 데이터 파일 경로")
    output_path: str = Field(description="출력 파일 경로 (.csv, .json, .xlsx, .parquet, .tsv)")
    sheet_name: str = Field(default="Sheet1", description="Excel 출력 시 시트 이름. 기본값 'Sheet1'.")

@tool(args_schema=FileConverterInput)
def file_converter(input_path: str, output_path: str, sheet_name: str = "Sheet1") -> str:
    """데이터 파일을 다른 포맷으로 변환합니다.

    지원 형식: CSV ↔ JSON ↔ Excel (.xlsx) ↔ Parquet ↔ TSV ↔ JSONL
    확장자를 기준으로 입출력 형식을 자동 감지합니다.

    Args:
        input_path: 입력 파일 경로.
        output_path: 출력 파일 경로 (확장자로 형식 결정).
        sheet_name: Excel 출력 시 시트 이름.

    Returns:
        변환 결과 메시지 (행 수, 파일 크기).
    """
    try:
        import pandas as pd
        df = _load_dataframe(input_path)
    except Exception as e:
        return f"FileConverter 로드 오류: {str(e)}"

    abs_output = os.path.abspath(os.path.expanduser(output_path))
    os.makedirs(os.path.dirname(abs_output), exist_ok=True)
    ext = Path(abs_output).suffix.lower()

    try:
        if ext == '.csv':
            df.to_csv(abs_output, index=False, encoding='utf-8-sig')
        elif ext == '.tsv':
            df.to_csv(abs_output, index=False, sep='\t', encoding='utf-8-sig')
        elif ext == '.json':
            df.to_json(abs_output, orient='records', force_ascii=False, indent=2)
        elif ext == '.jsonl':
            df.to_json(abs_output, orient='records', lines=True, force_ascii=False)
        elif ext in ('.xlsx', '.xls'):
            df.to_excel(abs_output, index=False, sheet_name=sheet_name)
        elif ext == '.parquet':
            df.to_parquet(abs_output, engine='pyarrow', compression='snappy')
        else:
            return f"FileConverter 오류: 지원하지 않는 출력 형식 '{ext}'"

        file_size = os.path.getsize(abs_output)
        input_ext = Path(input_path).suffix.upper()
        return (
            f"✅ 변환 완료\n"
            f"- **입력**: `{input_path}` ({input_ext})\n"
            f"- **출력**: `{abs_output}` ({ext.upper()})\n"
            f"- **행 수**: {len(df):,}\n"
            f"- **열 수**: {len(df.columns)}\n"
            f"- **크기**: {file_size/1024:.1f} KB"
        )

    except Exception as e:
        return f"FileConverter 저장 오류: {type(e).__name__}: {str(e)}"


# =============================================================================
# Tool 5: excel_writer — 수식/서식 기반 전문 Excel 보고서 생성
# =============================================================================

class ExcelWriterInput(BaseModel):
    data_source: str = Field(description="원본 데이터 파일 경로 (JSON, CSV, Excel 등)")
    output_path: str = Field(description="출력 Excel 파일 경로 (.xlsx)")
    excel_code: str = Field(description="openpyxl 기반 Excel 생성 코드. 'df'로 데이터 접근, 'wb'(Workbook)와 'ws'(active sheet)로 워크북 접근.")

@tool(args_schema=ExcelWriterInput)
def excel_writer(data_source: str, output_path: str, excel_code: str) -> str:
    """전문적인 수식·서식이 포함된 Excel 보고서를 생성합니다.

    openpyxl을 사용하여 수식, 조건부 서식, 차트가 포함된 프로페셔널 Excel을 생성합니다.
    코드에서 'df'로 원본 데이터, 'wb'(Workbook)와 'ws'(active sheet)로 워크북에 접근합니다.

    ⚠️ 핵심 규칙:
    - 수식 사용 필수, 하드코딩 금지 (=SUM(B2:B9), not Python 계산값)
    - XLOOKUP/XMATCH 사용 금지 → INDEX/MATCH 사용
    - 상세 가이드: file_read('skills/analyst/xlsx_guide.md')

    Args:
        data_source: 원본 데이터 파일 경로.
        output_path: 출력 .xlsx 파일 경로.
        excel_code: openpyxl 기반 Excel 생성 코드.

    Returns:
        생성된 Excel 파일 경로와 메타 정보.
    """
    try:
        import pandas as pd
        import numpy as np
        df = _load_dataframe(data_source)
    except Exception as e:
        return f"ExcelWriter 로드 오류: {str(e)}"

    abs_output = os.path.abspath(os.path.expanduser(output_path))
    os.makedirs(os.path.dirname(abs_output), exist_ok=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active

        exec_globals = {
            '__builtins__': {
                'len': len, 'range': range, 'enumerate': enumerate, 'zip': zip,
                'sorted': sorted, 'list': list, 'dict': dict, 'str': str,
                'int': int, 'float': float, 'min': min, 'max': max,
                'sum': sum, 'abs': abs, 'round': round, 'print': print,
                'True': True, 'False': False, 'None': None,
            },
            'pd': pd, 'np': np, 'df': df,
            'wb': wb, 'ws': ws,
            'Workbook': Workbook,
            'Font': Font, 'PatternFill': PatternFill, 'Alignment': Alignment,
            'Border': Border, 'Side': Side, 'numbers': numbers,
            'BarChart': BarChart, 'LineChart': LineChart, 'PieChart': PieChart,
            'Reference': Reference, 'get_column_letter': get_column_letter,
        }

        exec(excel_code, exec_globals, {})

        wb.save(abs_output)
        file_size = os.path.getsize(abs_output)
        sheet_names = wb.sheetnames

        return (
            f"✅ Excel 보고서 생성 완료\n"
            f"- **파일**: `{abs_output}`\n"
            f"- **크기**: {file_size/1024:.1f} KB\n"
            f"- **시트**: {', '.join(sheet_names)}\n"
            f"- **데이터**: {len(df)}행 from `{data_source}`"
        )

    except Exception as e:
        return f"ExcelWriter 실행 오류: {type(e).__name__}: {str(e)}\n{traceback.format_exc()[-500:]}"


# =============================================================================
# Tool 6: html_report — 인터랙티브 HTML 대시보드 생성
# =============================================================================

class HtmlReportInput(BaseModel):
    report_code: str = Field(description="HTML 보고서 생성 Python 코드. 'html_content' 변수에 최종 HTML 문자열을 저장합니다.")
    output_path: str = Field(description="출력 HTML 파일 경로")
    title: str = Field(default="분석 보고서", description="보고서 제목")

@tool(args_schema=HtmlReportInput)
def html_report(report_code: str, output_path: str, title: str = "분석 보고서") -> str:
    """분석 결과를 종합한 인터랙티브 HTML 대시보드를 생성합니다.

    Chart.js, Mermaid CDN을 포함한 단일 HTML 파일을 생성합니다.
    코드에서 'html_content' 변수에 최종 HTML 문자열을 저장하세요.
    'title' 변수로 보고서 제목에 접근 가능합니다.

    ⚠️ 디자인 가이드: file_read('skills/analyst/design_tokens.md')

    Args:
        report_code: HTML 생성 Python 코드.
        output_path: 출력 HTML 파일 경로.
        title: 보고서 제목.

    Returns:
        생성된 HTML 파일 경로와 메타 정보.
    """
    abs_output = os.path.abspath(os.path.expanduser(output_path))
    os.makedirs(os.path.dirname(abs_output), exist_ok=True)

    try:
        import pandas as pd
        import numpy as np

        # import 허용 모듈 제한
        _ALLOWED_MODULES = {'json', 'math', 'datetime', 'collections', 're', 'statistics'}
        _original_import = __builtins__.__import__ if hasattr(__builtins__, '__import__') else __import__

        def _safe_import(name, *args, **kwargs):
            if name in _ALLOWED_MODULES:
                return _original_import(name, *args, **kwargs)
            raise ImportError(f"모듈 '{name}'은 보안상 import할 수 없습니다. 허용: {_ALLOWED_MODULES}")

        exec_globals = {
            '__builtins__': {
                'len': len, 'range': range, 'enumerate': enumerate, 'zip': zip,
                'sorted': sorted, 'reversed': reversed, 'list': list, 'dict': dict,
                'set': set, 'tuple': tuple, 'str': str, 'int': int, 'float': float,
                'bool': bool, 'min': min, 'max': max, 'sum': sum, 'abs': abs,
                'round': round, 'isinstance': isinstance, 'type': type, 'print': print,
                'True': True, 'False': False, 'None': None,
                'open': open,
                '__import__': _safe_import,
            },
            'pd': pd,
            'np': np,
            'json': json,
            'title': title,
            'Path': Path,
            'os': os,
            '_load_dataframe': _load_dataframe,
        }
        exec_locals = {}

        exec(report_code, exec_globals, exec_locals)

        html_content = exec_locals.get('html_content', None)
        if html_content is None:
            return "HtmlReport 오류: 'html_content' 변수에 HTML 문자열을 저장하세요."

        with open(abs_output, 'w', encoding='utf-8') as f:
            f.write(html_content)

        file_size = os.path.getsize(abs_output)
        return (
            f"✅ HTML 보고서 생성 완료\n"
            f"- **파일**: `{abs_output}`\n"
            f"- **크기**: {file_size/1024:.1f} KB\n"
            f"- **제목**: {title}\n"
            f"브라우저에서 열거나 Chainlit에서 cl.File()로 공유하세요."
        )

    except Exception as e:
        return f"HtmlReport 실행 오류: {type(e).__name__}: {str(e)}\n{traceback.format_exc()[-500:]}"
