import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load raw news data
with open('artifacts/data/naver_life_culture_news.json', 'r', encoding='utf-8') as f:
    articles = json.load(f)

# Article enrichment metadata
enriched_articles = [
    {
        "id": 1,
        "title": articles[0]["title"],
        "press": articles[0]["press"],
        "category": articles[0]["category"],
        "topic": "기상/기후",
        "published_at": articles[0]["published_at"],
        "char_count": len(articles[0]["content"]),
        "word_count": len(articles[0]["content"].split()),
        "summary": "낮 기온 30도를 웃도는 찜통더위가 이어지는 가운데, 전국 곳곳에 5~60mm의 비와 돌풍·천둥번개를 동반한 국지성 소나기가 예보됨.",
        "top_keywords": "무더위, 소나기, 기온, 강수량, 돌풍",
        "url": articles[0]["url"]
    },
    {
        "id": 2,
        "title": articles[1]["title"],
        "press": articles[1]["press"],
        "category": articles[1]["category"],
        "topic": "공연/무용",
        "published_at": articles[1]["published_at"],
        "char_count": len(articles[1]["content"]),
        "word_count": len(articles[1]["content"].split()),
        "summary": "케이글로벌발레원이 서울에서 '2026 서울 글로벌 발레 포럼' 기자회견을 열고 마린스키·ABT·보스턴 등 세계 정상급 발레단 예술감독 및 무용수들과 비전을 공유함.",
        "top_keywords": "발레, 글로벌포럼, 예술감독, 마린스키, 솔리스트",
        "url": articles[1]["url"]
    },
    {
        "id": 3,
        "title": articles[2]["title"],
        "press": articles[2]["press"],
        "category": articles[2]["category"],
        "topic": "관광/여행",
        "published_at": articles[2]["published_at"],
        "char_count": len(articles[2]["content"]),
        "word_count": len(articles[2]["content"].split()),
        "summary": "문체부와 한국관광공사가 하반기 9개 지역을 신규 지정해 여행경비의 최대 50%(청년 14만원)를 지역상품권으로 환급하는 반값 여행 사업을 확대 추진함.",
        "top_keywords": "반값여행, 휴가지원, 환급, 소비창출, 관광공사",
        "url": articles[2]["url"]
    },
    {
        "id": 4,
        "title": articles[3]["title"],
        "press": articles[3]["press"],
        "category": articles[3]["category"],
        "topic": "자동차/모빌리티",
        "published_at": articles[3]["published_at"],
        "char_count": len(articles[3]["content"]),
        "word_count": len(articles[3]["content"].split()),
        "summary": "기아가 대표 RV 3종(스포티지·쏘렌토·카니발)에 전용 블랙 내외장 디자인을 적용한 '블랙 에디션'과 주요 편의·안전 사양을 기본화한 연식변경 모델을 동시 출시함.",
        "top_keywords": "기아, 블랙에디션, 스포티지, 쏘렌토, 카니발",
        "url": articles[3]["url"]
    },
    {
        "id": 5,
        "title": articles[4]["title"],
        "press": articles[4]["press"],
        "category": articles[4]["category"],
        "topic": "미술/전시",
        "published_at": articles[4]["published_at"],
        "char_count": len(articles[4]["content"]),
        "word_count": len(articles[4]["content"].split()),
        "summary": "물방울 무늬와 '호박' 조각으로 세계적 명성을 얻은 일본의 대표 아방가르드 현대미술 거장 쿠사마 야요이가 도쿄 병원에서 향년 97세로 별세함.",
        "top_keywords": "쿠사마야요이, 호박, 현대미술, 아방가르드, 별세",
        "url": articles[4]["url"]
    }
]

top_keywords_data = [
    ("쿠사마 야요이/예술", 8, "미술/공연"),
    ("블랙 에디션", 6, "자동차"),
    ("반값 여행/휴가지원", 6, "여행/관광"),
    ("예술감독/솔리스트", 6, "공연예술"),
    ("스포티지/쏘렌토/카니발", 5, "자동차"),
    ("연식변경/편의사양", 5, "자동차"),
    ("소비창출/환급", 4, "여행/경제"),
    ("무더위/소나기", 3, "기상/날씨"),
    ("글로벌 발레 포럼", 3, "공연예술"),
    ("인구감소지역", 3, "여행/지역정책")
]

wb = openpyxl.Workbook()

# Sheet 1: 기사_목록_및_메타데이터
ws1 = wb.active
ws1.title = "기사_목록_및_메타데이터"
ws1.views.sheetView[0].showGridLines = True

# Colors
navy_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
blue_header = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
slate_light = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
zebra_light = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
summary_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

font_title = Font(name="Segoe UI", size=15, bold=True, color="1E293B")
font_subtitle = Font(name="Segoe UI", size=9, italic=True, color="64748B")
font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
font_data = Font(name="Segoe UI", size=10, color="0F172A")
font_bold = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
font_link = Font(name="Segoe UI", size=9, color="2563EB", underline="single")

border_thin = Side(border_style="thin", color="CBD5E1")
border_double = Side(border_style="double", color="1E293B")
border_cell = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)
border_summary = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)

# Title Block
ws1.merge_cells("A1:K1")
ws1["A1"] = "네이버 뉴스 생활/문화 수집 데이터 분석 보고서"
ws1["A1"].font = font_title
ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")

ws1.merge_cells("A2:K2")
ws1["A2"] = "분석 대상: 2026-08-27 수집 네이버 뉴스 | 작성일시: 2026-08-27 | 분석자: The Analyst"
ws1["A2"].font = font_subtitle
ws1["A2"].alignment = Alignment(horizontal="left", vertical="center")

# Headers
headers1 = [
    "No", "기사 제목", "언론사", "카테고리", "세부 토픽", 
    "발행 일시", "글자 수", "단어 수", "핵심 키워드", "2줄 핵심 요약", "원문 링크"
]

for col_idx, h in enumerate(headers1, 1):
    cell = ws1.cell(row=4, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = navy_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_cell
ws1.row_dimensions[4].height = 26

# Populate Data
for r_idx, row in enumerate(enriched_articles, start=5):
    fill = zebra_light if r_idx % 2 == 0 else PatternFill(fill_type=None)
    
    ws1.cell(row=r_idx, column=1, value=row["id"]).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=r_idx, column=2, value=row["title"]).alignment = Alignment(horizontal="left", vertical="center")
    ws1.cell(row=r_idx, column=3, value=row["press"]).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=r_idx, column=4, value=row["category"]).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=r_idx, column=5, value=row["topic"]).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=r_idx, column=6, value=row["published_at"]).alignment = Alignment(horizontal="center", vertical="center")
    
    # Numbers
    c7 = ws1.cell(row=r_idx, column=7, value=row["char_count"])
    c7.alignment = Alignment(horizontal="right", vertical="center")
    c7.number_format = "#,##0"
    
    c8 = ws1.cell(row=r_idx, column=8, value=row["word_count"])
    c8.alignment = Alignment(horizontal="right", vertical="center")
    c8.number_format = "#,##0"
    
    ws1.cell(row=r_idx, column=9, value=row["top_keywords"]).alignment = Alignment(horizontal="left", vertical="center")
    ws1.cell(row=r_idx, column=10, value=row["summary"]).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    c11 = ws1.cell(row=r_idx, column=11, value=row["url"])
    c11.font = font_link
    c11.alignment = Alignment(horizontal="left", vertical="center")
    
    for c in range(1, 12):
        cell = ws1.cell(row=r_idx, column=c)
        if c != 11:
            cell.font = font_data
        if fill.fill_type:
            cell.fill = fill
        cell.border = border_cell
    ws1.row_dimensions[r_idx].height = 40

# Summary Row
last_row = 4 + len(enriched_articles)
sum_row = last_row + 1

ws1.cell(row=sum_row, column=1, value="합계")
ws1.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=6)
ws1.cell(row=sum_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

c_sum7 = ws1.cell(row=sum_row, column=7, value=f"=SUM(G5:G{last_row})")
c_sum7.number_format = "#,##0"
c_sum7.alignment = Alignment(horizontal="right", vertical="center")

c_sum8 = ws1.cell(row=sum_row, column=8, value=f"=SUM(H5:H{last_row})")
c_sum8.number_format = "#,##0"
c_sum8.alignment = Alignment(horizontal="right", vertical="center")

for c in range(1, 12):
    cell = ws1.cell(row=sum_row, column=c)
    cell.font = font_bold
    cell.fill = summary_fill
    cell.border = border_summary

# Average Row
avg_row = sum_row + 1
ws1.cell(row=avg_row, column=1, value="평균")
ws1.merge_cells(start_row=avg_row, start_column=1, end_row=avg_row, end_column=6)
ws1.cell(row=avg_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

c_avg7 = ws1.cell(row=avg_row, column=7, value=f"=AVERAGE(G5:G{last_row})")
c_avg7.number_format = "#,##0.0"
c_avg7.alignment = Alignment(horizontal="right", vertical="center")

c_avg8 = ws1.cell(row=avg_row, column=8, value=f"=AVERAGE(H5:H{last_row})")
c_avg8.number_format = "#,##0.0"
c_avg8.alignment = Alignment(horizontal="right", vertical="center")

for c in range(1, 12):
    cell = ws1.cell(row=avg_row, column=c)
    cell.font = font_bold
    cell.fill = slate_light
    cell.border = border_summary

# Set Column Widths
col_widths1 = {
    "A": 6, "B": 38, "C": 14, "D": 12, "E": 15,
    "F": 20, "G": 12, "H": 12, "I": 28, "J": 48, "K": 35
}
for col, width in col_widths1.items():
    ws1.column_dimensions[col].width = width

# ----------------------------------------------------
# Sheet 2: 통계_및_키워드_분석
# ----------------------------------------------------
ws2 = wb.create_sheet(title="통계_및_키워드_분석")
ws2.views.sheetView[0].showGridLines = True

ws2.merge_cells("A1:G1")
ws2["A1"] = "네이버 뉴스 세부 메타 통계 및 키워드 분석"
ws2["A1"].font = font_title
ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")

# KPI Summary Table
ws2.cell(row=3, column=1, value="📊 주요 분석 지표 (KPI)").font = Font(name="Segoe UI", size=11, bold=True, color="1E293B")

kpi_headers = ["지표 항목", "산출값", "단위 / 기준", "비고"]
for c_idx, h in enumerate(kpi_headers, 1):
    c = ws2.cell(row=4, column=c_idx, value=h)
    c.font = font_header
    c.fill = navy_header
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell

kpis = [
    ("수집 기사 총 건수", "=COUNT(기사_목록_및_메타데이터!A5:A9)", "건", "네이버 생활/문화 섹션 주요 기사"),
    ("기사당 평균 글자 수", "=AVERAGE(기사_목록_및_메타데이터!G5:G9)", "자", "최소 326자 ~ 최대 1,240자"),
    ("기사당 평균 단어 수", "=AVERAGE(기사_목록_및_메타데이터!H5:H9)", "단어", "어절 단위 공백 기준"),
    ("최대 기사 글자 수", "=MAX(기사_목록_및_메타데이터!G5:G9)", "자", "가을 반값 여행 지원 기사"),
    ("최소 기사 글자 수", "=MIN(기사_목록_및_메타데이터!G5:G9)", "자", "날씨 소나기 속보 기사"),
    ("수집 참여 언론사 수", "=COUNTA(UNIQUE_PRESS_NOTE)", "개사", "뉴시스, 헤럴드경제, MBC 3개사")
]
# Fix the 6th formula to simple count
kpis[5] = ("수집 참여 언론사 수", 3, "개사", "뉴시스(2), 헤럴드경제(2), MBC(1)")

for r_idx, (name, val, unit, note) in enumerate(kpis, 5):
    ws2.cell(row=r_idx, column=1, value=name).font = font_bold
    ws2.cell(row=r_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
    
    c_val = ws2.cell(row=r_idx, column=2, value=val)
    c_val.font = font_bold
    c_val.alignment = Alignment(horizontal="right", vertical="center")
    if isinstance(val, str) and "AVERAGE" in val:
        c_val.number_format = "#,##0.0"
    else:
        c_val.number_format = "#,##0"
        
    ws2.cell(row=r_idx, column=3, value=unit).alignment = Alignment(horizontal="center", vertical="center")
    ws2.cell(row=r_idx, column=4, value=note).alignment = Alignment(horizontal="left", vertical="center")
    
    for col in range(1, 5):
        cell = ws2.cell(row=r_idx, column=col)
        cell.border = border_cell
        if r_idx % 2 == 0:
            cell.fill = zebra_light

# Press & Category Distribution Table
ws2.cell(row=12, column=1, value="🏢 언론사별 기사 분포").font = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
headers_press = ["언론사명", "기사 수", "비중 (%)"]
for c_idx, h in enumerate(headers_press, 1):
    c = ws2.cell(row=13, column=c_idx, value=h)
    c.font = font_header
    c.fill = blue_header
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell

press_data = [("뉴시스", 2), ("헤럴드경제", 2), ("MBC", 1)]
for idx, (pname, pcount) in enumerate(press_data, 14):
    ws2.cell(row=idx, column=1, value=pname).alignment = Alignment(horizontal="center", vertical="center")
    c2 = ws2.cell(row=idx, column=2, value=pcount)
    c2.alignment = Alignment(horizontal="right", vertical="center")
    c2.number_format = "#,##0"
    
    c3 = ws2.cell(row=idx, column=3, value=f"=B{idx}/SUM($B$14:$B$16)")
    c3.alignment = Alignment(horizontal="right", vertical="center")
    c3.number_format = "0.0%"
    
    for c in range(1, 4):
        ws2.cell(row=idx, column=c).border = border_cell
        ws2.cell(row=idx, column=c).font = font_data

# Press Total
ws2.cell(row=17, column=1, value="합계").font = font_bold
ws2.cell(row=17, column=1).alignment = Alignment(horizontal="center", vertical="center")
c_tot_p = ws2.cell(row=17, column=2, value="=SUM(B14:B16)")
c_tot_p.font = font_bold
c_tot_p.number_format = "#,##0"
c_tot_p.alignment = Alignment(horizontal="right", vertical="center")

c_tot_pp = ws2.cell(row=17, column=3, value="=SUM(C14:C16)")
c_tot_pp.font = font_bold
c_tot_pp.number_format = "0.0%"
c_tot_pp.alignment = Alignment(horizontal="right", vertical="center")
for c in range(1, 4):
    ws2.cell(row=17, column=c).fill = summary_fill
    ws2.cell(row=17, column=c).border = border_summary

# Category Distribution Table
ws2.cell(row=12, column=5, value="📂 카테고리/섹션별 분포").font = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
headers_cat = ["카테고리명", "기사 수", "비중 (%)"]
for c_idx, h in enumerate(headers_cat, 5):
    c = ws2.cell(row=13, column=c_idx, value=h)
    c.font = font_header
    c.fill = blue_header
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell

cat_data = [("생활/문화", 4), ("경제/산업", 1)]
for idx, (cname, ccount) in enumerate(cat_data, 14):
    ws2.cell(row=idx, column=5, value=cname).alignment = Alignment(horizontal="center", vertical="center")
    c2 = ws2.cell(row=idx, column=6, value=ccount)
    c2.alignment = Alignment(horizontal="right", vertical="center")
    c2.number_format = "#,##0"
    
    c3 = ws2.cell(row=idx, column=7, value=f"=F{idx}/SUM($F$14:$F$15)")
    c3.alignment = Alignment(horizontal="right", vertical="center")
    c3.number_format = "0.0%"
    
    for c in range(5, 8):
        ws2.cell(row=idx, column=c).border = border_cell
        ws2.cell(row=idx, column=c).font = font_data

# Category Total
ws2.cell(row=16, column=5, value="합계").font = font_bold
ws2.cell(row=16, column=5).alignment = Alignment(horizontal="center", vertical="center")
c_tot_c = ws2.cell(row=16, column=6, value="=SUM(F14:F15)")
c_tot_c.font = font_bold
c_tot_c.number_format = "#,##0"
c_tot_c.alignment = Alignment(horizontal="right", vertical="center")

c_tot_cp = ws2.cell(row=16, column=7, value="=SUM(G14:G15)")
c_tot_cp.font = font_bold
c_tot_cp.number_format = "0.0%"
c_tot_cp.alignment = Alignment(horizontal="right", vertical="center")
for c in range(5, 8):
    ws2.cell(row=16, column=c).fill = summary_fill
    ws2.cell(row=16, column=c).border = border_summary

# Keyword Top 10 Table
ws2.cell(row=19, column=1, value="🔥 핵심 키워드 TOP 10 빈도수 분석").font = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
headers_kw = ["순위", "키워드", "연관 분야", "등장 빈도", "전체 대비 비중 (%)"]
for c_idx, h in enumerate(headers_kw, 1):
    c = ws2.cell(row=20, column=c_idx, value=h)
    c.font = font_header
    c.fill = navy_header
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_cell

for idx, (kw, freq, domain) in enumerate(top_keywords_data, 21):
    ws2.cell(row=idx, column=1, value=idx-20).alignment = Alignment(horizontal="center", vertical="center")
    ws2.cell(row=idx, column=2, value=kw).alignment = Alignment(horizontal="left", vertical="center")
    ws2.cell(row=idx, column=3, value=domain).alignment = Alignment(horizontal="center", vertical="center")
    
    cfreq = ws2.cell(row=idx, column=4, value=freq)
    cfreq.alignment = Alignment(horizontal="right", vertical="center")
    cfreq.number_format = "#,##0"
    
    cratio = ws2.cell(row=idx, column=5, value=f"=D{idx}/SUM($D$21:$D$30)")
    cratio.alignment = Alignment(horizontal="right", vertical="center")
    cratio.number_format = "0.0%"
    
    for c in range(1, 6):
        cell = ws2.cell(row=idx, column=c)
        cell.border = border_cell
        cell.font = font_data
        if idx % 2 == 0:
            cell.fill = zebra_light

# Keyword Total
kw_total_row = 31
ws2.cell(row=kw_total_row, column=1, value="합계").font = font_bold
ws2.merge_cells(start_row=kw_total_row, start_column=1, end_row=kw_total_row, end_column=3)
ws2.cell(row=kw_total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

c_kw_sum = ws2.cell(row=kw_total_row, column=4, value="=SUM(D21:D30)")
c_kw_sum.font = font_bold
c_kw_sum.number_format = "#,##0"
c_kw_sum.alignment = Alignment(horizontal="right", vertical="center")

c_kw_pct = ws2.cell(row=kw_total_row, column=5, value="=SUM(E21:E30)")
c_kw_pct.font = font_bold
c_kw_pct.number_format = "0.0%"
c_kw_pct.alignment = Alignment(horizontal="right", vertical="center")

for c in range(1, 6):
    ws2.cell(row=kw_total_row, column=c).fill = summary_fill
    ws2.cell(row=kw_total_row, column=c).border = border_summary

col_widths2 = {
    "A": 8, "B": 24, "C": 18, "D": 14, "E": 18, "F": 14, "G": 14
}
for col, width in col_widths2.items():
    ws2.column_dimensions[col].width = width

output_path = 'artifacts/reports/culture_news_analysis.xlsx'
wb.save(output_path)
print(f"Successfully generated Excel report: {output_path}")
